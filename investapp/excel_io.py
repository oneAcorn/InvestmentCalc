"""Excel 导出 / 导入（可往返）。"""
from __future__ import annotations

from datetime import date, datetime

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

from .models import (BASE_CURRENCY, CURRENCIES, INVESTMENT_TYPES, Portfolio,
                     Investment, Transaction, parse_date)

HEADER_FILL = PatternFill("solid", fgColor="305496")
HEADER_FONT = Font(color="FFFFFF", bold=True)

INV_HEADERS = ["名称", "类型", "币种", "当前市值(条目币种)", "当前汇率", "备注"]
TX_HEADERS = ["投资名称", "日期", "类型", "金额(条目币种)", "汇率",
              "当日持仓市值(条目币种,可选)", "备注"]

NOTE_LINES = [
    "投资条目导出/导入说明",
    "",
    f"基础货币: {BASE_CURRENCY}",
    "",
    "投资 sheet 每行一个投资条目；交易 sheet 通过“投资名称”关联。",
    "",
    "类型取值: " + "、".join(INVESTMENT_TYPES),
    "币种取值: " + "、".join(CURRENCIES),
    "交易类型取值: 买入/卖出/存入/取出/申购/赎回/分红/利息",
    "（买入/申购/存入=投入资金；卖出/赎回/取出/分红/利息=收回资金）",
    "金额以条目币种记录(正数)；汇率 = 1 条目币种折合多少人民币(人民币条目填 1)。",
    "当日持仓市值可选，用于精确计算 TWRR(时间加权收益率)。",
]


def _style_header(ws, ncols: int) -> None:
    for c in range(1, ncols + 1):
        cell = ws.cell(row=1, column=c)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")


def export_to_excel(path: str, portfolio: Portfolio) -> None:
    wb = Workbook()

    ws_inv = wb.active
    ws_inv.title = "投资"
    ws_inv.append(INV_HEADERS)
    for inv in portfolio.investments:
        ws_inv.append([inv.name, inv.itype, inv.currency,
                       inv.current_value, inv.current_fx, inv.note])
    _style_header(ws_inv, len(INV_HEADERS))

    ws_tx = wb.create_sheet("交易")
    ws_tx.append(TX_HEADERS)
    for inv in portfolio.investments:
        for t in sorted(inv.transactions, key=lambda x: x.date):
            ws_tx.append([inv.name, t.date, t.type, t.amount, t.fx,
                          t.holding_value, t.note])
    _style_header(ws_tx, len(TX_HEADERS))
    for row in ws_tx.iter_rows(min_row=2, min_col=2, max_col=2):
        for cell in row:
            cell.number_format = "yyyy-mm-dd"

    ws_note = wb.create_sheet("说明")
    for i, line in enumerate(NOTE_LINES, start=1):
        ws_note.cell(row=i, column=1, value=line)
    ws_note.column_dimensions["A"].width = 80

    for ws, widths in ((ws_inv, [22, 14, 8, 20, 12, 30]),
                       (ws_tx, [22, 12, 10, 18, 10, 24, 30])):
        for i, w in enumerate(widths, start=1):
            ws.column_dimensions[get_column_letter(i)].width = w

    wb.save(path)


def import_from_excel(path: str) -> Portfolio:
    wb = load_workbook(path, data_only=True)
    if "投资" not in wb.sheetnames:
        raise ValueError("Excel 中缺少“投资”工作表")

    ws_inv = wb["投资"]
    rows_inv = list(ws_inv.iter_rows(min_row=2, values_only=True))
    rows_inv = [r for r in rows_inv if r and r[0] not in (None, "")]

    investments = []
    names = set()
    for r in rows_inv:
        name = str(r[0]).strip()
        itype = str(r[1]).strip() if r[1] else "理财"
        currency = str(r[2]).strip() if r[2] else "CNY"
        if name in names:
            raise ValueError(f"投资名称重复: {name}")
        if itype not in INVESTMENT_TYPES:
            raise ValueError(f"未知投资类型: {itype}")
        if currency not in CURRENCIES:
            raise ValueError(f"未知币种: {currency}")
        names.add(name)
        investments.append(Investment(
            name=name,
            itype=itype,
            currency=currency,
            current_value=float(r[3]) if r[3] is not None else 0.0,
            current_fx=float(r[4]) if r[4] else 1.0,
            note=str(r[5]).strip() if len(r) > 5 and r[5] else "",
        ))

    if "交易" in wb.sheetnames:
        ws_tx = wb["交易"]
        for r in ws_tx.iter_rows(min_row=2, values_only=True):
            if not r or r[0] in (None, ""):
                continue
            inv = _find_by_name(investments, str(r[0]).strip())
            if inv is None:
                raise ValueError(f"交易引用了不存在的投资: {r[0]}")
            tdate = _to_date(r[1])
            ttype = str(r[2]).strip() if r[2] else "买入"
            amount = float(r[3]) if r[3] is not None else 0.0
            fx = float(r[4]) if r[4] else 1.0
            hval = float(r[5]) if r[5] is not None else None
            note = str(r[6]).strip() if len(r) > 6 and r[6] else ""
            inv.transactions.append(Transaction(
                date=tdate, type=ttype, amount=amount, fx=fx,
                holding_value=hval, note=note,
            ))

    return Portfolio(investments)


def _find_by_name(investments, name):
    for inv in investments:
        if inv.name == name:
            return inv
    return None


def _to_date(v) -> date:
    if isinstance(v, datetime):
        return v.date()
    if isinstance(v, date):
        return v
    if isinstance(v, str):
        for fmt in ("%Y-%m-%d", "%Y/%m/%d", "%Y.%m.%d"):
            try:
                return datetime.strptime(v.strip(), fmt).date()
            except ValueError:
                continue
    raise ValueError(f"无法解析日期: {v!r}")
